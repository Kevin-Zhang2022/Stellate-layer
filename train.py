import datetime
import time

import numpy as np
from cla.Data_set import Data_set as ds
from torch.utils.data import DataLoader
from cla.Net import Net
import torch
import torch.nn as nn
from cla.Show import Show as show
from cla.Data_process import Data_process as dp
import matplotlib.pyplot as plt
from openpyxl import Workbook
import multiprocessing as processing
import logging
import time
from datetime import datetime
import math
import sys
import shutil



# def seed_torch(seed=10):
#     np.random.seed(seed)
#     os.environ['PYTHONHASHSEED'] = str(seed)
#     np.random.seed(seed)
#     torch.manual_seed(seed)
#     # torch.backends.cudnn.benchmark = False
#     # torch.backends.cudnn.deterministic = True
# seed_torch()

# def setup_seed(seed):
#     torch.manual_seed(seed)
#     torch.cuda.manual_seed_all(seed)
#     np.random.seed(seed)
#     random.seed(seed)
#     torch.backends.cudnn.deterministic = True
#
#
# # 设置随机数种子
# setup_seed(20)


# check hyper parameters
# bandwidth_factor   hidden batch size channel
#
#
trials = 5  # dft 55 10
epochs = 9  # dft 920 12
epochs_rep = 3
sc_step_size = 3  # dft 3
max_duration = 4  # dft 5


# trials = 1  # 10
# reps = 1  # 5 3
# epochs = 1  # 20 12
# sc_step_size = 1  # 5 3


gtf_fs = 10000  # sampling frequency fs hz
gtf_fr=[20,5000]  # frequency range hz
device = torch.device('cuda:0' if torch.cuda.is_available() else 'cpu')
data_type = torch.float32
loss = nn.CrossEntropyLoss()
train_batchsize_dic = {'us8k': 10,'CWRU':40,'esc10':40}
test_batchsize_dic = {'us8k': 10,'CWRU':40,'esc10':40}


input_list = [100,200,300,500]  # list k

task_list = ['us8k','CWRU','esc10']
out_dic = {'us8k': 10,'CWRU':10,'esc10':10}

sigma_b_lis=[20,15,12,10,8,5]  # list i
sf_lis=[1,0.8,0.5,0.3]  # list j

ihc_win = 50
gtf_bf = 0.05
an_uth = 0.05
model = 'mh'
task = 'us8k'  # task_list = [ 'us8k','CWRU','esc10']

for k in [0]: # for input_list
    for i in range(len(sigma_b_lis)):  # sigma_b_lis
        for j in range(len(sf_lis)):  # bushy_range_list  one_list1
            for t in range(1):
                print(datetime.now())
                out_features = out_dic[task]
                train_batchsize = train_batchsize_dic[task]
                test_batchsize = test_batchsize_dic[task]

                in_features = input_list[k]  # input_list = [100,200,300,500,800,1000]
                sigma_b = sigma_b_lis[i]  # sigma_b_lis[i]
                sf = sf_lis[j]  # sf_lis[j]

                sigma_lis = [sigma_b, sigma_b * sf, sigma_b * sf * sf]

                src='data/'+task
                # tar='data/'+task+f'_{k:d}_{i:d}_{j:d}'
                # shutil.copytree(src,tar)


                dp.create_datalist(path2audio=src + '/audio')
                train_dataset = ds(data_list_path=src+ '/train_list.csv', max_duration=max_duration, sample_rate=gtf_fs)
                test_dataset = ds(data_list_path=src + '/test_list.csv', max_duration=max_duration, sample_rate=gtf_fs)
                train_loader = DataLoader(dataset=train_dataset, shuffle=True, batch_size=train_batchsize)
                test_loader = DataLoader(dataset=test_dataset, shuffle=True, batch_size=test_batchsize)
                net = Net(in_features=in_features, out_features=out_features,
                          gtf_bf=gtf_bf, gtf_fr=gtf_fr, gtf_fs=gtf_fs,
                          ihc_win=ihc_win,
                          an_uth=an_uth,
                          steli_std=sigma_lis)
                optimizer = torch.optim.Adam(net.parameters(), lr=0.1, betas=(0.9, 0.999))
                scheduler = torch.optim.lr_scheduler.StepLR(optimizer, step_size=sc_step_size, gamma=0.1)

                print('finish dataset setting',f'k:{k:d} i:{i:d} j:{j:d} t:{t:d}')

                # logo='finish dataset setting '+f'k:{k:d} i:{i:d} j:{j:d} t:{t:d}'
                # sys.stdout.write(logo + '\n')
                sys.stdout.flush()

                process = 'train'
                for epoch in range(epochs):
                    # print('after epoch')
                    sys.stdout.flush()
                    for batch_id, (audio, label) in enumerate(train_loader):
                        # print('after batch')
                        # sys.stdout.flush()
                        start=time.time()
                        data = dp.normalize(audio)
                        out = net(data)
                        end=time.time()
                        print(end-start)
                        loss_val = show.loss_val(loss, out, label)
                        acc = show.acc(out, label)
                        # a = net.fc0.weight.grad
                        optimizer.zero_grad()
                        loss_val.backward(retain_graph=True)
                        optimizer.step()
                        print('p:', process, 'k:', k, 'i:', i, 'j:', j, 't:', t, 'e:', epoch, 'acc:', f'{acc:.2f}')
                        sys.stdout.flush()
                    scheduler.step()
                # torch.save(net,'../net_for_figtab/average cc net.pkl')

                process = 'ttrain'
                acc_rec = []
                for epoch in range(epochs_rep):
                    for batch_id, (audio, label) in enumerate(train_loader):
                        data = dp.normalize(audio)
                        out = net(data)
                        # torch.save(net,'../net_for_figtab/bushy_ic bushy_range3.pkl')
                        loss_val = show.loss_val(loss, out, label)
                        acc = show.acc(out, label)
                        acc_rec.append(acc)
                        print('p:', process, 'k:', k, 'i:', i, 'j:', j, 't:', t, 'e:', epoch, 'acc:', f'{acc:.2f}')
                        sys.stdout.flush()
                train_acc = np.mean(acc_rec)

                process = 'ttest'
                acc_rec = []
                snr_an_rec = []
                snr_osf_b_rec = []
                snr_osf_m_rec = []
                snr_osf_n_rec = []
                for epoch in range(epochs_rep):
                    for batch_id, (audio, label) in enumerate(test_loader):
                        data = dp.normalize(audio)
                        out = net(data)
                        loss_val = show.loss_val(loss, out, label)
                        acc = show.acc(out, label)
                        acc_rec.append(acc)
                        snr_an_rec.append(net.snr_an)
                        snr_osf_b_rec.append(net.snr_osf_b)
                        snr_osf_m_rec.append(net.snr_osf_m)
                        snr_osf_n_rec.append(net.snr_osf_n)

                        print('p:', process, 'k:', k, 'i:', i, 'j:', j, 't:', t, 'e:', epoch, 'acc:', f'{acc:.2f}')
                        sys.stdout.flush()
                test_acc = np.mean(acc_rec)
                snr_an = np.array(torch.mean(torch.tensor(snr_an_rec)))
                snr_osf_b = np.array(torch.mean(torch.tensor(snr_osf_b_rec)))
                snr_osf_m = np.array(torch.mean(torch.tensor(snr_osf_m_rec)))
                snr_osf_n = np.array(torch.mean(torch.tensor(snr_osf_n_rec)))

                # save dta
                print('save')
                sys.stdout.flush()
                wb = Workbook()
                sheet = wb['Sheet']
                sheet.title = 'result'

                wb['result'].cell(1, 1, f'{np.mean(train_acc):.4f}')
                wb['result'].cell(1, 2, f'{np.mean(test_acc):.4f}')
                wb['result'].cell(1, 3, f'{np.mean(snr_an):.4f}')
                wb['result'].cell(1, 4, f'{np.mean(snr_osf_b):.4f}')
                wb['result'].cell(1, 5, f'{np.mean(snr_osf_m):.4f}')
                wb['result'].cell(1, 6, f'{np.mean(snr_osf_n):.4f}')
                wb.save('tab/' + task + '-' + str(k) + '-' + str(i) + '-' + str(j) + '-' + str(t) + '.xlsx')
                # shutil.rmtree(tar)


print('end')

    # single_train(0,2,3,0)
    # a=10






