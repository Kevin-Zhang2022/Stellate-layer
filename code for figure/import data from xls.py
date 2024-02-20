import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook


acc_train = np.zeros([4,6,6])
acc_test = np.zeros([4,6,6])
snr_b = np.zeros([4,6,6])
snr_a = np.zeros([4,6,6])

# input_list = [100,200,300,500]  # list k 4
# sigma_b_lis=[20,15,12,10,8,5]  # list i 6
# sf_lis=[1,0.8,0.5,0.3]  # list j 4

for k in range(4):
    for i in range(6):
        for j in range(4):
            filename= f'../tab/us8k-{k:d}-{i:d}-{j:d}-{0:d}.xlsx'
            wb = load_workbook(filename)
            # acc_train[k,i,j]= wb['result'].cell(1,1).value
            acc_test[k,i,j] = wb['result'].cell(1,2).value
            snr_b[k,i,j]= wb['result'].cell(1,3).value
            snr_a[k,i,j] = wb['result'].cell(1,6).value


# filename=f'../tab/CWRU-on/temp.xlsx'
# wb = load_workbook(filename)
wb = Workbook()
sheet0 = wb.create_sheet('sheet0')
# sheet1 = wb.create_sheet('sheet1')
# sheet2 = wb.create_sheet('sheet2')
# sheet3 = wb.create_sheet('sheet3')

# i_list = [np.arange(2,8),np.arange(2,8)+6,np.arange(2,8)+12]
# j_list= [np.arange(3,7),np.arange(3,7)+5]
for k in range(4):
    for i in range(6):
        for j in range(4):
            # wb[f'sheet0'].cell(k*24+i*6+j+1, 1, f'{acc_train[k,i,j]:.4f}')
            wb[f'sheet0'].cell(k*24+i*4+j+1, 2, f'{acc_test[k,i,j]:.4f}')
            wb[f'sheet0'].cell(k*24+i*4+j+1, 3, f'{snr_b[k,i,j]:.4f}')
            wb[f'sheet0'].cell(k*24+i*4+j+1, 4, f'{snr_a[k,i,j]:.4f}')

wb.save('../tab/all/temp.xlsx')