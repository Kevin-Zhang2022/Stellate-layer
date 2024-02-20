import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook
from scipy import stats

acc_train = np.zeros([4,6,6])
acc_test = np.zeros([4,6,6])
snr_b = np.zeros([4,6,6])
snr_a = np.zeros([4,6,6])

# input_list = [100,200,300,500]  # list k 4
# sigma_b_lis=[20,15,12,10,8,5]  # list i 6
# sf_lis=[1,0.8,0.5,0.3]  # list j 4

K_i=[[],[]]
A_i=[[],[]]
filename = f'../tab/us8k.xlsx'
wb = load_workbook(filename)
for i in range(2,98):
    # acc_train[k,i,j]= wb['result'].cell(1,1).value
    K_i[0].append(wb['final results'].cell(i,1).value)
    A_i[0].append(wb['final results'].cell(i,2).value)

filename = f'../tab/esc10.xlsx'
wb = load_workbook(filename)
for i in range(2,98):
    # acc_train[k,i,j]= wb['result'].cell(1,1).value
    K_i[1].append(wb['final results'].cell(i,1).value)
    A_i[1].append(wb['final results'].cell(i,2).value)


fig = plt.figure(figsize=(8,6))
fig.show()
fontsize=8
# x=[0.3,0.5,0.8,1,2,3,5,8,10,20]
# title_lis=['(a)','(b)','(c)','(d)','(e)','(f)']
color_list=['r','k','gold','g','royalblue','deeppink','purple']
label_list=['r=0.49 \np=4.5E-7','r=0.25 \np=1.3E-2']
linestyle_list = ['solid','dotted','dashed','dashdot',(0, (1, 1)),(0, (5, 10)),(0, (3, 10, 1, 10, 1, 10))]

axe1 = fig.add_subplot(1, 2, 1)
axe2 = fig.add_subplot(1, 2, 2)
axe1.scatter(K_i[0], A_i[0],label=label_list[0])
axe2.scatter(K_i[1], A_i[1],label=label_list[1])

axe1.set_ylim([-0.05,0.3])
axe2.set_ylim([-0.05,0.3])
axe1.set_xlim([0,4.5])
axe2.set_xlim([0,4.5])
# axe2.set_ylim([0.2,0.9])
axe1.grid('on')
axe2.grid('on')
axe1.legend(loc='lower right',fontsize=fontsize*1.5)
axe2.legend(loc='lower right',fontsize=fontsize*1.5)
# axe1.set_xlabel('$K_{\mathrm{sn}}$',fontsize=fontsize*1.5)
axe1.set_xlabel('$K_{i}$',fontsize=fontsize*1.5)
axe1.set_title('(a) US8K',y=-0.15)
axe1.set_ylabel('$A_i$',fontsize=fontsize*1.5)

axe2.set_xlabel('$K_{i}$',fontsize=fontsize*1.5)
axe2.set_title('(b) ESC10',y=-0.15)
axe2.set_ylabel('$A_i$',fontsize=fontsize*1.5)

# r indicates cc
# p indicates t_US8K
# The p-value roughly indicates the probability of an uncorrelated system i.e. significance test
CC_US8K, t_US8K = stats.pearsonr(K_i[0], A_i[0])
CC_ESC10, t_ESC10 = stats.pearsonr(K_i[1], A_i[1])

# axe1.plot([min(K_i[0]), min(K_i[0])+2], [min(A_i[0]), min(A_i[0])+2*CC_US8K])  # 绘制直线


fig.subplots_adjust(left=0.10,bottom=0.12,right=0.98,top=0.98,wspace=0.27,hspace=0.4)
fig.show()
filename = '../fig/E/E Correlation Analysis/E Correlation Analysis.pdf'
fig.savefig(filename,bbox_inches='tight')

# filename=f'../tab/CWRU-on/temp.xlsx'
# wb = load_workbook(filename)
# plt.figure()
# plt.scatter(K_i[0],A_i[0])
# wb = Workbook()




# sheet0 = wb.create_sheet('sheet0')
# # sheet1 = wb.create_sheet('sheet1')
# # sheet2 = wb.create_sheet('sheet2')
# # sheet3 = wb.create_sheet('sheet3')
#
# # i_list = [np.arange(2,8),np.arange(2,8)+6,np.arange(2,8)+12]
# # j_list= [np.arange(3,7),np.arange(3,7)+5]
# for k in range(4):
#     for i in range(6):
#         for j in range(4):
#             # wb[f'sheet0'].cell(k*24+i*6+j+1, 1, f'{acc_train[k,i,j]:.4f}')
#             wb[f'sheet0'].cell(k*24+i*4+j+1, 2, f'{acc_test[k,i,j]:.4f}')
#             wb[f'sheet0'].cell(k*24+i*4+j+1, 3, f'{snr_b[k,i,j]:.4f}')
#             wb[f'sheet0'].cell(k*24+i*4+j+1, 4, f'{snr_a[k,i,j]:.4f}')
#
# wb.save('../tab/all/temp.xlsx')